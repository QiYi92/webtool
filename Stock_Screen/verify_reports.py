from __future__ import annotations

import argparse
import logging
import shutil
import time
from datetime import datetime, time as dt_time, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import screen_bowl_shape as screener


BASE_DIR = Path("verificate_data")
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
CACHE_DIR = BASE_DIR / "cache"
RUN_RECORD_SHEET = "运行记录"
MARKET_CLOSE_CONFIRM_TIME = dt_time(15, 30)
KLINE_LOOKBACK_COUNT = 20
OUTPUT_COLUMNS = [
    "股票代码",
    "股票名称",
    "报告完成时间",
    "报告日期",
    "验证交易日",
    "报告日收盘价",
    "验证日开盘价",
    "验证日最高价",
    "验证日最低价",
    "验证日收盘价",
    "最高点涨幅%",
    "最低点跌幅%",
    "收盘涨跌幅%",
    "收盘方向",
    "备注",
]


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)


def ensure_directories() -> None:
    """创建验证模块需要的输入、输出和缓存目录。"""
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    CACHE_DIR.mkdir(parents=True, exist_ok=True)


def configure_verification_cache() -> None:
    """让复用的行情函数把缓存写入验证模块自己的缓存目录。"""
    screener.CACHE_DIR = CACHE_DIR


def cleanup_cache_dir() -> None:
    """运行结束后清理验证缓存目录。"""
    if not CACHE_DIR.exists():
        return

    removed_count = 0
    for path in CACHE_DIR.iterdir():
        if path.is_file():
            path.unlink()
            removed_count += 1
        elif path.is_dir():
            shutil.rmtree(path)
            removed_count += 1

    logging.info("已清理验证缓存文件 %s 个", removed_count)


def normalize_stock_code(value: Any) -> str:
    """把 Excel 中的股票代码统一格式化为 6 位字符串。"""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text.zfill(6)


def parse_report_finished_at(value: Any) -> pd.Timestamp | None:
    """解析运行记录中的完成时间。"""
    if value is None or pd.isna(value):
        return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed


def list_input_reports() -> list[Path]:
    """获取输入目录中的待验证报告。"""
    ensure_directories()
    return sorted(
        path
        for path in INPUT_DIR.glob("*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    )


def read_run_record(report_path: Path) -> tuple[pd.Timestamp, pd.DataFrame]:
    """读取报告中的运行完成时间和命中股票清单。"""
    wb = load_workbook(report_path, data_only=True, read_only=True)
    if RUN_RECORD_SHEET not in wb.sheetnames:
        raise RuntimeError(f"缺少 {RUN_RECORD_SHEET} 工作表")

    ws = wb[RUN_RECORD_SHEET]
    finished_at = parse_report_finished_at(ws["B2"].value)
    if finished_at is None:
        raise RuntimeError("运行记录 B2 缺少有效的运行完成时间")

    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        code = row[0] if len(row) >= 1 else None
        name = row[1] if len(row) >= 2 else None
        if code is None and name is None:
            continue
        if code is None:
            continue
        rows.append(
            {
                "股票代码": normalize_stock_code(code),
                "股票名称": "" if name is None else str(name).strip(),
            }
        )

    return finished_at, pd.DataFrame(rows, columns=["股票代码", "股票名称"])


def next_weekday(date_value: pd.Timestamp) -> pd.Timestamp:
    """在无法从 K 线确认时，估算下一个非周末日期用于未收市提示。"""
    candidate = date_value.normalize() + pd.Timedelta(days=1)
    while candidate.weekday() >= 5:
        candidate += pd.Timedelta(days=1)
    return candidate


def is_after_market_close(trade_date: pd.Timestamp, now: datetime | None = None) -> bool:
    """判断当前是否已经过目标交易日 15:30。"""
    current = pd.Timestamp(now or datetime.now())
    target_close = pd.Timestamp.combine(trade_date.date(), MARKET_CLOSE_CONFIRM_TIME)
    return current >= target_close


def fetch_stock_kline(symbol: str) -> pd.DataFrame:
    """拉取用于验证的最近 K 线。"""
    return screener.fetch_recent_k(symbol, KLINE_LOOKBACK_COUNT)


def find_base_and_verify_rows(
    k_df: pd.DataFrame,
    report_date: pd.Timestamp,
) -> tuple[pd.Series | None, pd.Series | None]:
    """找到报告日基准 K 线，以及基准日后的第一根验证 K 线。"""
    if k_df.empty:
        return None, None

    df = k_df.copy()
    df["日期"] = pd.to_datetime(df["日期"], errors="coerce").dt.normalize()
    df = df.dropna(subset=["日期"]).sort_values("日期").reset_index(drop=True)

    base_candidates = df[df["日期"] <= report_date.normalize()]
    if base_candidates.empty:
        return None, None

    base_row = base_candidates.iloc[-1]
    verify_candidates = df[df["日期"] > base_row["日期"]]
    if verify_candidates.empty:
        return base_row, None

    return base_row, verify_candidates.iloc[0]


def detect_report_verification_day(
    stocks: pd.DataFrame,
    report_date: pd.Timestamp,
) -> tuple[pd.Timestamp | None, str | None]:
    """用第一只有 K 线的股票确认报告的目标验证交易日。"""
    if stocks.empty:
        return None, None

    errors: list[str] = []
    for _, row in stocks.iterrows():
        symbol = row["股票代码"]
        try:
            k_df = fetch_stock_kline(symbol)
            base_row, verify_row = find_base_and_verify_rows(k_df, report_date)
            if verify_row is not None:
                return pd.Timestamp(verify_row["日期"]).normalize(), None
            if base_row is not None:
                estimated_day = next_weekday(pd.Timestamp(base_row["日期"]))
                if not is_after_market_close(estimated_day):
                    return None, f"目标验证日 {estimated_day:%Y-%m-%d} 尚未到 15:30 收市确认时间"
                errors.append(f"{symbol}: 缺少基准日后的验证 K 线")
        except Exception as exc:
            errors.append(f"{symbol}: {exc}")
            continue

    estimated_day = next_weekday(report_date)
    if not is_after_market_close(estimated_day):
        return None, f"目标验证日 {estimated_day:%Y-%m-%d} 尚未到 15:30 收市确认时间"
    return None, "无法确认目标验证交易日，可能是休市、行情尚未更新或接口缺失: " + " | ".join(errors[:5])


def build_empty_result_row(
    stock: pd.Series,
    finished_at: pd.Timestamp,
    report_date: pd.Timestamp,
    remark: str,
) -> dict[str, Any]:
    """构造行情失败时的占位结果行。"""
    return {
        "股票代码": stock["股票代码"],
        "股票名称": stock["股票名称"],
        "报告完成时间": finished_at.strftime("%Y-%m-%d %H:%M:%S"),
        "报告日期": report_date.strftime("%Y-%m-%d"),
        "验证交易日": "",
        "报告日收盘价": "",
        "验证日开盘价": "",
        "验证日最高价": "",
        "验证日最低价": "",
        "验证日收盘价": "",
        "最高点涨幅%": "",
        "最低点跌幅%": "",
        "收盘涨跌幅%": "",
        "收盘方向": "",
        "备注": remark,
    }


def build_verified_result_row(
    stock: pd.Series,
    finished_at: pd.Timestamp,
    report_date: pd.Timestamp,
    base_row: pd.Series,
    verify_row: pd.Series,
) -> dict[str, Any]:
    """根据报告日和验证日 K 线计算一只股票的验证结果。"""
    base_close = float(base_row["收盘"])
    verify_open = float(verify_row["开盘"])
    verify_high = float(verify_row["最高"])
    verify_low = float(verify_row["最低"])
    verify_close = float(verify_row["收盘"])
    high_pct = (verify_high / base_close - 1) * 100
    low_pct = (verify_low / base_close - 1) * 100
    close_pct = (verify_close / base_close - 1) * 100

    if close_pct > 0:
        direction = "上涨"
    elif close_pct < 0:
        direction = "下跌"
    else:
        direction = "持平"

    return {
        "股票代码": stock["股票代码"],
        "股票名称": stock["股票名称"],
        "报告完成时间": finished_at.strftime("%Y-%m-%d %H:%M:%S"),
        "报告日期": report_date.strftime("%Y-%m-%d"),
        "验证交易日": pd.Timestamp(verify_row["日期"]).strftime("%Y-%m-%d"),
        "报告日收盘价": round(base_close, 3),
        "验证日开盘价": round(verify_open, 3),
        "验证日最高价": round(verify_high, 3),
        "验证日最低价": round(verify_low, 3),
        "验证日收盘价": round(verify_close, 3),
        "最高点涨幅%": round(high_pct, 2),
        "最低点跌幅%": round(low_pct, 2),
        "收盘涨跌幅%": round(close_pct, 2),
        "收盘方向": direction,
        "备注": "",
    }


def verify_one_report(report_path: Path, sleep_seconds: float) -> pd.DataFrame | None:
    """验证一份筛选报告，返回验证结果表。"""
    finished_at, stocks = read_run_record(report_path)
    report_date = finished_at.normalize()
    logging.info("读取报告 %s，运行完成时间: %s，命中股票数: %s", report_path.name, finished_at, len(stocks))

    if stocks.empty:
        logging.info("%s 没有命中股票，将生成空验证结果", report_path.name)
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    verify_day, stop_reason = detect_report_verification_day(stocks, report_date)
    if verify_day is None:
        logging.warning("%s 暂不验证: %s", report_path.name, stop_reason)
        return None

    if not is_after_market_close(verify_day):
        logging.warning("%s 暂不验证: 目标验证日 %s 尚未到 15:30", report_path.name, verify_day.strftime("%Y-%m-%d"))
        return None

    rows: list[dict[str, Any]] = []
    for index, stock in stocks.iterrows():
        symbol = stock["股票代码"]
        try:
            k_df = fetch_stock_kline(symbol)
            base_row, verify_row = find_base_and_verify_rows(k_df, report_date)
            if base_row is None:
                rows.append(build_empty_result_row(stock, finished_at, report_date, "缺少报告日或之前的基准 K 线"))
            elif verify_row is None:
                rows.append(build_empty_result_row(stock, finished_at, report_date, "缺少基准日后的验证 K 线"))
            else:
                rows.append(build_verified_result_row(stock, finished_at, report_date, base_row, verify_row))
        except Exception as exc:
            rows.append(build_empty_result_row(stock, finished_at, report_date, f"行情获取失败: {exc}"))

        if index % 20 == 0:
            logging.info("验证进度 %s: %s/%s", report_path.name, index + 1, len(stocks))
        time.sleep(sleep_seconds)

    return pd.DataFrame(rows, columns=OUTPUT_COLUMNS)


def build_output_path(report_path: Path) -> Path:
    """生成验证结果输出路径。"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return OUTPUT_DIR / f"verify_{report_path.stem}_{timestamp}.xlsx"


def style_result_sheet(ws) -> None:
    """设置验证结果表样式。"""
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill

    ws.freeze_panes = "A2"
    for column_index in range(1, ws.max_column + 1):
        letter = get_column_letter(column_index)
        ws.column_dimensions[letter].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["O"].width = 34


def export_verification_excel(df: pd.DataFrame, output_path: Path) -> None:
    """导出单份报告对应的验证 Excel。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "验证结果"

    for column_index, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        ws.cell(row=1, column=column_index, value=column_name)

    for row_index, (_, row) in enumerate(df.iterrows(), start=2):
        for column_index, column_name in enumerate(OUTPUT_COLUMNS, start=1):
            ws.cell(row=row_index, column=column_index, value=row.get(column_name, ""))

    style_result_sheet(ws)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="验证筛选报告命中股票在下一个交易日的走势")
    parser.add_argument(
        "--input-dir",
        default=str(INPUT_DIR),
        help="待验证报告目录，默认 verificate_data/input",
    )
    parser.add_argument(
        "--output-dir",
        default=str(OUTPUT_DIR),
        help="验证结果输出目录，默认 verificate_data/output",
    )
    parser.add_argument(
        "--sleep",
        type=float,
        default=0.05,
        help="每只股票请求后的暂停秒数",
    )
    parser.add_argument(
        "--use-proxy",
        action="store_true",
        help="使用当前终端代理；默认关闭代理以减少行情接口连接失败",
    )
    return parser.parse_args()


def main() -> None:
    global INPUT_DIR, OUTPUT_DIR

    args = parse_args()
    INPUT_DIR = Path(args.input_dir)
    OUTPUT_DIR = Path(args.output_dir)

    ensure_directories()
    configure_verification_cache()
    screener.configure_network(use_proxy=args.use_proxy)

    reports = list_input_reports()
    if not reports:
        logging.info("没有找到待验证报告，请把 .xlsx 文件放到 %s", INPUT_DIR)
        cleanup_cache_dir()
        return

    generated_count = 0
    try:
        for report_path in reports:
            try:
                result_df = verify_one_report(report_path, sleep_seconds=args.sleep)
                if result_df is None:
                    continue
                output_path = build_output_path(report_path)
                export_verification_excel(result_df, output_path)
                generated_count += 1
                logging.info("验证结果已保存: %s", output_path)
            except Exception as exc:
                logging.warning("跳过报告 %s: %s", report_path.name, exc)
    finally:
        cleanup_cache_dir()

    logging.info("验证完成，生成 %s 份验证结果", generated_count)


if __name__ == "__main__":
    main()
