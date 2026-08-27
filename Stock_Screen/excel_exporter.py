from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


UNKNOWN_SECTOR = "未分类"


def normalize_stock_code(value: Any) -> str:
    """把股票代码显示成 6 位字符串。"""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text.zfill(6)


def get_sorted_result_df(df: pd.DataFrame) -> pd.DataFrame:
    """按板块和趋势指标排序，缺失字段时自动降级。"""
    if df.empty:
        return df

    sort_columns = [
        col
        for col in ["板块", "最近一年涨跌幅", "最近15日涨跌幅", "总市值"]
        if col in df.columns
    ]
    ascending = [True] + [False] * (len(sort_columns) - 1)
    return df.sort_values(sort_columns, ascending=ascending) if sort_columns else df


def write_sector_block(
    ws,
    sector_name: str,
    group: pd.DataFrame,
    start_col: int,
) -> None:
    """在工作表中写入一个板块的两列表格。"""
    code_col = start_col
    name_col = start_col + 1

    ws.merge_cells(start_row=1, start_column=code_col, end_row=1, end_column=name_col)
    title_cell = ws.cell(row=1, column=code_col, value=sector_name)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=2, column=code_col, value="股票代码")
    ws.cell(row=2, column=name_col, value="股票名称")

    for row_index, (_, row) in enumerate(group.iterrows(), start=3):
        ws.cell(row=row_index, column=code_col, value=normalize_stock_code(row["股票代码"]))
        ws.cell(row=row_index, column=name_col, value=str(row["股票名称"]))


def style_worksheet(ws) -> None:
    """统一设置工作表样式。"""
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row in (1, 2):
                cell.font = Font(bold=True, size=12 if cell.row == 2 else 14)
                cell.fill = header_fill

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 14 if col_idx % 2 == 1 else 18


def write_sector_summary_sheet(ws, df: pd.DataFrame) -> None:
    """按板块把股票代码和名称写入指定工作表。"""
    if df.empty:
        write_sector_block(
            ws=ws,
            sector_name=UNKNOWN_SECTOR,
            group=pd.DataFrame(columns=["股票代码", "股票名称"]),
            start_col=1,
        )
        style_worksheet(ws)
        return

    sorted_df = get_sorted_result_df(df)
    start_col = 1
    for sector, group in sorted_df.groupby("板块", dropna=False):
        sector_name = str(sector).strip() if str(sector).strip() else UNKNOWN_SECTOR
        write_sector_block(ws, sector_name, group, start_col)
        start_col += 2

    style_worksheet(ws)


def write_bowl_stage_sheets(wb: Workbook, df: pd.DataFrame) -> None:
    """按碗型阶段拆分工作表，便于直接查看不同启动类型。"""
    stage_column = "碗型阶段"
    stages = ("右侧萌芽", "早期启动")

    for stage in stages:
        ws = wb.create_sheet(stage)
        if stage_column not in df.columns:
            stage_df = pd.DataFrame(columns=df.columns)
        else:
            stage_df = df[df[stage_column] == stage]
        write_sector_summary_sheet(ws, stage_df)


def write_run_record_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    started_at: str | None = None,
    finished_at: str | None = None,
) -> None:
    """写入本次运行记录：时间戳和命中股票清单。"""
    ws = wb.create_sheet("运行记录")

    ws.cell(row=1, column=1, value="运行开始时间")
    ws.cell(row=1, column=2, value=started_at or "")
    ws.cell(row=2, column=1, value="运行完成时间")
    ws.cell(row=2, column=2, value=finished_at or "")
    ws.cell(row=3, column=1, value="命中数量")
    ws.cell(row=3, column=2, value=len(df))

    header_row = 5
    ws.cell(row=header_row, column=1, value="股票代码")
    ws.cell(row=header_row, column=2, value="股票名称")

    if not df.empty:
        sorted_df = get_sorted_result_df(df)
        for row_index, (_, row) in enumerate(sorted_df.iterrows(), start=header_row + 1):
            ws.cell(row=row_index, column=1, value=normalize_stock_code(row["股票代码"]))
            ws.cell(row=row_index, column=2, value=str(row["股票名称"]))

    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row in (1, 2, 3, header_row):
                cell.font = Font(bold=True)
                cell.fill = header_fill

    ws.freeze_panes = "A6"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22


def export_sector_summary_excel(
    df: pd.DataFrame,
    output_path: str | Path,
    started_at: str | None = None,
    finished_at: str | None = None,
) -> None:
    """导出板块汇总 Excel：每个板块占两列，并附带本次运行记录。"""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "板块汇总"

    write_sector_summary_sheet(ws, df)
    write_bowl_stage_sheets(wb, df)
    write_run_record_sheet(wb, df, started_at=started_at, finished_at=finished_at)
    wb.save(output)
